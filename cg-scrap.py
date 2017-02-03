import openpyxl
from time import sleep

import requests 
from bs4 import BeautifulSoup

def is_number(s):
    
    try:
        float(s)
        return True
    except ValueError:
        return False

def connect(fname, url_to_scrape):
   
    r = requests.get(url_to_scrape, verify = False)
    soup = BeautifulSoup(r.text, "html.parser")
    with open(fname, "w") as text_file:
        text_file.write("{}".format(soup))
    with open(fname) as f:
        content = f.readlines()
    return content




def find_cg_individual(roll_num, content = ''):
    
    count=0
    if content == '':
        url_to_scrape = 'https://erp.iitkgp.ernet.in/StudentPerformance/view_performance.jsp?rollno=' + str(roll_num)
        fname = "Output.txt"
        content = connect(fname, url_to_scrape)
    
    if len(content) < 50:
        return -1
    else:
        for line in content:
            
            if line.find("CGPA") != -1 and line[4] != "<" and is_number(line[31:35]):
            	
            	if count==1:
                	return float(line[31:35])
                count=1
        return -1

if __name__ == '__main__':


	wb = openpyxl.load_workbook('cg.xlsx')
	sheet = wb.get_sheet_by_name('Sheet1')
	n=sheet.max_row+1
	for i in range(1,n):
		sleep(0.5)
		roll= sheet.cell(row=i, column=1).value
		sheet['B'+str(i)]=find_cg_individual(roll,"")
	wb.save('cg.xlsx')
